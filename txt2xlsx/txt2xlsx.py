#!/usr/bin/python3
# -*- coding: UTF-8 -*-

import sys
import openpyxl
import os
import re


def is_subject(line):
	matchObj = re.match('^\d+\. ', line)
	if(matchObj):
		return True
	else:
		return False

def is_option(option):
	print("is: %s" % option)
	matchObj = re.match('^\s+[A-E]\.\s', option)
	if(matchObj):
		print(re.split('\. ', option, maxsplit=1))
		return True
	else:
		return False


def get_context(line):
	return re.split('\. ', line, maxsplit=1)[1]


def get_option(line):
	return re.split('\. ', line, maxsplit=1)[1]

def save_subject(ws, num, context, A, B, C, D, E):
	print('num: %d, context: %s' % (num, context) )
	row = num + 1
	ws.cell(row, column=1).value = num
	ws.cell(row, column=2).value = context
	ws.cell(row, column=3).value = A
	ws.cell(row, column=4).value = B
	ws.cell(row, column=5).value = C
	ws.cell(row, column=6).value = D
	ws.cell(row, column=7).value = E


if( len(sys.argv) != 3 ):
	print('Usage: ./txt2xlsx.py txt_file xlsx_file')
	exit()

in_file = str(sys.argv[1])

(filepath, tempfilename) = os.path.split(in_file)
(exam_name, extension) = os.path.splitext(tempfilename)

print('Gen exam paper: %s' % exam_name)
wb = openpyxl.workbook.Workbook()
ws = wb.create_sheet(exam_name, 0)

tableTitle = ['Number', 'Promt', 'A', 'B', 'C', 'D', 'E','Answer','Picture','Type','Period(1-9)','Theme','Point 1','Point 2','Point 3', \
		'Logic','Difficulty (1-6)']
for col in range(len(tableTitle)):
	c = col + 1
	ws.cell(row=1, column=c).value = tableTitle[col]

out_file = str(sys.argv[2])
print('Transfer %s to %s ...' % (in_file, out_file) )

try:
	f = open(in_file)
	line = f.readline()
	num = 0
	while line:
		print(line)
		if(is_subject(line)):
			num = num + 1
			context = 'N/A'
			A = 'N/A'
			B = 'N/A'
			C = 'N/A'
			D = 'N/A'
			E = 'N/A'
			context = get_context(line)
			option = f.readline() #try A
			if(option.strip()==''):
				print("This is a blank string, read again onec")
				option = f.readline() #avoid A
			if(is_option(option)):
				A = get_option(option)	#save A
				option = f.readline() #try B
				if(is_option(option)):
					B = get_option(option)	#save B
					option = f.readline() #try C
					if(is_option(option)):
						C = get_option(option)	#save C
						option = f.readline() #try D
						if(is_option(option)):
							D = get_option(option)	#save D
							option = f.readline() #try E
							if(is_option(option)):
								E = get_option(option)	#save E

			save_subject(ws, num, context, A, B, C, D, E)

		line = f.readline()
	f.close()
	wb.save(filename=out_file)

	print('done.')
except Exception as e:
	print(repr(e))



